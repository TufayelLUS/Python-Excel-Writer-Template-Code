#!/usr/bin/python3
# pip3 install openpyxl
import openpyxl
import os


def saveData(dataset, output_file):
    output_file = "{}.xlsx".format(output_file)
    fieldnames = ["Header 1", "Header 2", "Header 3",
                  "Header 4"]  # change your header list here
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()
    sheet = wb.active
    last_row = sheet.max_row
    if last_row == 1:
        idx = 1
        for fieldname in fieldnames:
            sheet.cell(row=1, column=idx).value = fieldname
            idx += 1
    idx = 1
    for data in dataset:
        sheet.cell(row=last_row+1, column=idx).value = data
        idx += 1
    wb.save(output_file)


if __name__ == "__main__":
    data = ["test", "test1", "test2", "test3"]
    saveData(data, "my_sheet")
